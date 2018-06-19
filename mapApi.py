# @Time    : 2018/6/10 10:13
# @Author  : Ghoul
# @FileName: mapApi.py
# @Software: PyCharm
# @Project： godMap

from urllib.parse import quote
from urllib import request
import json
from openpyxl import load_workbook

amap_web_key = '3ca4889dbb38a1984c64e9d720670895'
poi_search_url = "http://restapi.amap.com/v3/place/text"

line = 1

def getpois(cityname, keywords):
    i = 1
    poilist = []
    while True : #使用while循环不断分页获取数据
       result = getpoi_page(cityname, keywords, i)
       result = json.loads(result)  # 将字符串转换为json
       if result['count'] == '0':
           break
       hand(poilist, result)
       i = i + 1
    return poilist


def getpoi_page(cityname, keywords, page):
    print("b")
    req_url = poi_search_url + "?key=" + amap_web_key + '&extensions=all&types=' + str(keywords) + '&city=' + quote(cityname) + '&citylimit=true' + '&offset=25' + '&page=' + str(page) + '&output=json'
    data = ''
    with request.urlopen(req_url) as f:
        data = f.read()
        data = data.decode('utf-8')
    return data

def hand(poilist, result):
    print(result["infocode"])
    if result["infocode"] == str(10003):
        exit(-1)
    pois = result['pois']
    for i in range(len(pois)) :
        poilist.append(pois[i])


def save_to_excel_Municipalities(poi,city):
    file_name = "医院信息.xlsx"
    wb = load_workbook(filename=file_name)

    wb.create_sheet(title=str(city))
    ws = wb.get_sheet_by_name(city)
    ws["A1"] = "酒店名称"
    ws["B1"] = "详细地址"
    ws["C1"] = "电话号码"
    ws["D1"] = "官方网站"
    ws["E1"] = "服务类型"
    ws["F1"] = "经纬度"
    for index in range(len(poi)):
        # print(poi[index]["name"],poi[index]["type"],str(poi[index]["cityname"]) + str(poi[index]["adname"]) + str(poi[index]["address"]),poi[index]["tel"])
        ws["A"+ str(index + 2 )] = poi[index]["name"]

        ws["B"+ str(index + 2 )] = str(poi[index]["cityname"]) + str(poi[index]["adname"]) + str(poi[index]["address"])

        ws["E"+ str(index + 2 )] = poi[index]["type"]

        ws["F"+ str(index + 2 )] = poi[index]["location"]




        try:
            ws["C" + str(index + 2)] = poi[index]["tel"]
        except:
            ws["C" + str(index + 2)] = None

        try:
            ws["D" + str(index + 2)] = poi[index]["website"]
        except:
            ws["D" + str(index + 2)] = None


    wb.save(filename=file_name)

def save_to_excel_onesheet(poi,city):
    print(city)
    file_name = "医院信息.xlsx"
    wb = load_workbook(filename=file_name)

    global line
    # wb.create_sheet(title=str(city))
    # ws = wb.get_sheet_by_name(city)
    sheetnames = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheetnames[0])
    ws["A1"] = "医院名称"
    ws["B1"] = "详细地址"
    ws["C1"] = "电话号码"
    ws["D1"] = "官方网站"
    ws["E1"] = "省"
    ws["F1"] = "市"
    ws["G1"] = "工作类型"
    ws["H1"] = "经纬度"
    for index in range(len(poi)):
    # if index <= (line+len(poi)):
    # for index in range(line,line+len(poi)):
        # print(poi[index]["name"],poi[index]["type"],str(poi[index]["cityname"]) + str(poi[index]["adname"]) + str(poi[index]["address"]),poi[index]["tel"])
        ws["A"+ str(line  )] = poi[index]["name"]

        ws["B"+ str(line  )] = str(poi[index]["cityname"]) + str(poi[index]["adname"]) + str(poi[index]["address"])
        try:
            ws["C" + str(line )] = poi[index]["tel"]
        except:
            ws["C" + str(line )] = None

        try:
            ws["D" + str(line )] = poi[index]["website"]
        except:
            ws["D" + str(line )] = None

        ws["E" + str(line )] = poi[index]["pname"]
        ws["F" + str(line )] = poi[index]["cityname"]

        ws["G"+ str(line  )] = poi[index]["type"]

        ws["H"+ str(line )] = poi[index]["location"]

        line +=1




    wb.save(filename=file_name)


def cityInfom():
    city = []
    file_name = "城市信息.xlsx"
    wb = load_workbook(filename=file_name)
    sheetnames = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheetnames[0])
    max_row = ws.max_row
    for index in range(1,max_row+1):
        # alphbeat = chr(ord("A")+index)
        if ws["A" + str(index)].value == None:
            continue
        info =  ws["A" + str(index)].value
        # print(info)
        city.append(info)
    return city



if __name__ == '__main__':

    citys= cityInfom()
    print(citys)
    for city in citys:
        # pass
        poi = getpois(city,"090203")
        save_to_excel_Municipalities(poi,city)
    for city in citys:
        # pass
        poi = getpois(city,"090203")
        save_to_excel_onesheet(poi,city)